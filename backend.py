import os
import openpyxl
import requests
import re
from openpyxl.styles import PatternFill


def write_to_excel():
    try:
        wb = openpyxl.load_workbook('excel.xlsx')
    except:
        wb = openpyxl.Workbook()
        del wb['Sheet']
    if 'Архив' not in wb.sheetnames:
        wb.create_sheet('Архив')
    ws = wb['Архив']

    # для подсчёта серий
    three = 800
    four = 810
    fifth = 820

    dict_of_summ = {three: 0,
                    four: 0,
                    fifth: 0, }

    max1, max2, max3 = 0, 0, 0

    with open('test.txt', 'r', encoding='utf-8') as f:
        old_tirags = tuple(f.readlines())

    title = ('Тираж',) + tuple(x for x in range(1, 21)) + ('Сумма очков', '', 'Кол-во серий < 800', 'Кол-во серий < 810', 'Кол-во серий < 820')

    for i in enumerate(title):
        ws.cell(row=1, column=i[0] + 1).value = i[1]

    for index, one_tirage in enumerate(old_tirags[::-1]):
        ws.cell(row=index + 2, column=1).value = int(one_tirage[:one_tirage.find('; ')])
        one_tirage = one_tirage[one_tirage.find('; ') + 2:].split(', ')
        one_tirage[-1] = one_tirage[-1][:-1]
        one_tirage = tuple(int(x) for x in one_tirage)

        for k in enumerate(one_tirage):
            ws.cell(row=index + 2, column=k[0] + 2).value = k[1]
        else:
            one_summ = sum(one_tirage)
            ws.cell(row=index + 2, column=22).value = one_summ

            # серия
            if one_summ < three:
                if max1 >= dict_of_summ[three]:
                    max1 += 1
                    dict_of_summ[three] = max1
                else:
                    max1 += 1
            else:
                max1 = 0

            if one_summ < four:
                if max2 >= dict_of_summ[four]:
                    max2 += 1
                    dict_of_summ[four] = max2
                else:
                    max2 += 1
            else:
                max2 = 0

            if one_summ < fifth:
                if max3 >= dict_of_summ[fifth]:
                    max3 += 1
                    dict_of_summ[fifth] = max3
                else:
                    max3 += 1
            else:
                max3 = 0

    ws.cell(row=2, column=24).value = dict_of_summ.get(800)
    ws.cell(row=2, column=25).value = dict_of_summ.get(810)
    ws.cell(row=2, column=26).value = dict_of_summ.get(820)

    try:
        wb.save('excel.xlsx')
    except:
        wb.close()
        return 'Обновление не удалось, закройте файл и повторите.'
    else:
        wb.close()
        return True


def parse():
    # задаем начальные данные
    data = {
        'page': '1',
        'mode': 'date',
        'super': 'false',
        'from': '10.10.2015',
        'to': '12.04.2020',
    }
    first_url , second_url = 'https://www.stoloto.ru/keno/archive', 'https://www.stoloto.ru/draw-results/keno/load'
    request = requests.post(first_url)
    current_url = first_url

    old_tirags = list()
    with open('test.txt', 'r', encoding='utf-8') as f:
        old_tirags = tuple(f.readlines()[:-5])

    last_tirag = str()
    if len(old_tirags) > 1:
        last_tirag = old_tirags[-1]

    list_of_new_tirags = list()

    with open('test.txt', 'w', encoding='utf-8') as f:

        while request.status_code == 200:
            request = request.text
            for one_tirag in re.findall(r"/keno/archive/[^⚲]*</span>", request):
                date = re.search(r'\d{6}', one_tirag).group() + '; '
                if last_tirag.startswith(date):
                    f.write(''.join(old_tirags + tuple(list_of_new_tirags[::-1])))
                    print('Данные добавлены')
                    return
                else:
                    if current_url == first_url:
                        one_tirag = one_tirag[one_tirag.find('<div class="container cleared">'):]
                    else:
                        one_tirag = one_tirag[one_tirag.find('<div class=\\"container cleared\\">'):]
                    bals = re.findall(r'\d\d?', one_tirag)
                    list_of_new_tirags.append(date + ', '.join(bals) + '\n')
            if current_url == first_url:
                current_url = second_url
            else:
                data['page'] = str(int(data.get('page')) + 1)
            request = requests.post(current_url, data=data)
            print("Страница № " + data.get('page'))
        else:
            f.write(''.join(tuple(list_of_new_tirags[::-1])))
            print("Все данные считаны")


def sorting(amount):
    wb = openpyxl.load_workbook(filename='excel.xlsx')
    list_of_tirags = list()
    with open('test.txt', 'r', encoding='utf-8') as f:
        list_of_tirags = f.readlines()[::-1][:amount]

    if 'Отфильтрованные записи' not in wb.sheetnames:
        wb.create_sheet('Отфильтрованные записи')
    else:
        del wb['Отфильтрованные записи']
        wb.create_sheet('Отфильтрованные записи')
    ws = wb['Отфильтрованные записи']

    title = ('Тираж',) + tuple(x for x in range(1, 21)) + ('Сумма очков', 'MAX', 'MIN', 'Серия MAX', 'Серия MIN', 'Кол-во серий < 800', 'Кол-во серий < 810', 'Кол-во серий < 820')

    for i in enumerate(title):
        ws.cell(row=1, column=i[0] + 1).value = i[1]

    # для подсчёта серий
    one = 78
    two = 3
    three = 800
    four = 810
    fifth = 820

    dict_of_summ = {one: 0,
                    two: 0,
                    three: 0,
                    four: 0,
                    fifth: 0, }

    max1, max2, max3 = 0, 0, 0
    max4, max5 = 0, 0

    for index, one_tirage in enumerate(list_of_tirags):
        ws.cell(row=index + 2, column=1).value = int(one_tirage[:one_tirage.find('; ')])
        one_tirage = one_tirage[one_tirage.find('; ') + 2:].split(', ')
        one_tirage[-1] = one_tirage[-1][:-1]
        one_tirage = tuple(int(x) for x in one_tirage)

        max_ = max(one_tirage)
        min_ = min(one_tirage)
        for k in enumerate(one_tirage):
            ws.cell(row=index + 2, column=k[0] + 2).value = k[1]
            if k[1] == max_:
                color = '92D050'
                ws.cell(row=index + 2, column=k[0] + 2).fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
            elif k[1] == min_:
                color = 'FFD966'
                ws.cell(row=index + 2, column=k[0] + 2).fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
        else:
            ws.cell(row=index + 2, column=23).value = max_
            color = '92D050' if max_ > 78 else 'FFFFFF'
            ws.cell(row=index + 2, column=23).fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
            ws.cell(row=index + 2, column=24).value = min_
            color = 'FFD966' if min_ < 3 else 'FFFFFF'
            ws.cell(row=index + 2, column=24).fill = PatternFill(fill_type='solid', start_color=color, end_color=color)


            if max_ > one:
                if max4 >= dict_of_summ[one]:
                    max4 += 1
                    dict_of_summ[one] = max4
                else:
                    max4 += 1
            else:
                max4 = 0

            if min_ < two:
                if max5 >= dict_of_summ[two]:
                    max5 += 1
                    dict_of_summ[two] = max5
                else:
                    max5 += 1
            else:
                max5 = 0

            one_summ = sum(one_tirage)
            ws.cell(row=index + 2, column=22).value = one_summ

            # серия
            if one_summ < three:
                if max1 >= dict_of_summ[three]:
                    max1 += 1
                    dict_of_summ[three] = max1
                else:
                    max1 += 1
            else:
                max1 = 0

            if one_summ < four:
                if max2 >= dict_of_summ[four]:
                    max2 += 1
                    dict_of_summ[four] = max2
                else:
                    max2 += 1
            else:
                max2 = 0

            if one_summ < fifth:
                if max3 >= dict_of_summ[fifth]:
                    max3 += 1
                    dict_of_summ[fifth] = max3
                else:
                    max3 += 1
            else:
                max3 = 0

    ws.cell(row=2, column=25).value = dict_of_summ.get(one)
    ws.cell(row=2, column=25).fill = PatternFill(fill_type='solid', start_color='9BC2E6', end_color='9BC2E6')
    ws.cell(row=2, column=26).value = dict_of_summ.get(two)
    ws.cell(row=2, column=26).fill = PatternFill(fill_type='solid', start_color='9BC2E6', end_color='9BC2E6')
    ws.cell(row=2, column=27).value = dict_of_summ.get(three)
    ws.cell(row=2, column=27).fill = PatternFill(fill_type='solid', start_color='92D050', end_color='92D050')
    ws.cell(row=2, column=28).value = dict_of_summ.get(four)
    ws.cell(row=2, column=28).fill = PatternFill(fill_type='solid', start_color='92D050', end_color='92D050')
    ws.cell(row=2, column=29).value = dict_of_summ.get(fifth)
    ws.cell(row=2, column=29).fill = PatternFill(fill_type='solid', start_color='92D050', end_color='92D050')
    try:
        wb.save('excel.xlsx')
    except:
        wb.close()
        return 'Сортировка не удалась, закройте файл и повторите попытку.'
    else:
        wb.close()
        os.startfile('excel.xlsx')
        return True


if __name__ == '__main__':
    # sorting(1000)
    write_to_excel()
    # parse()
    pass